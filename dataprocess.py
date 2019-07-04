import re

from openpyxl.reader.excel import load_workbook
import json
from copy import deepcopy
from xpinyin import Pinyin
pin = Pinyin()

res = set()
file = open("D:\work_space\敏感词库\网站敏感词库\dictionaries\chinese_dictionary.txt", 'r', encoding='utf-8')
for line in file:
    res.add(line.strip())
    res.add(pin.get_pinyin(line.strip()).replace("-", ""))
file = open("D:\work_space\敏感词库\网站敏感词库\dictionaries\english_dictionary.txt", 'r', encoding='utf-8')
for line in file:
    res.add(line.strip())
    res.add(pin.get_pinyin(line.strip()).replace("-", ""))
file = open("D:\work_space\敏感词库\网站敏感词库\JasonYSU的敏感词库\key.txt", 'r', encoding='utf-8')
for line in file:
    for item in line.split("|"):
        res.add(item.strip())
        res.add(pin.get_pinyin(item.strip()).replace("-", ""))

file = open("D:\work_space\敏感词库\网站敏感词库\百度敏感词\百度过滤词.txt", 'r', encoding='gbk')
for line in file:
    res.add(line.strip())
    res.add(pin.get_pinyin(line.strip()).replace("-", ""))

file = open("D:\work_space\敏感词库\网站敏感词库\百度敏感词\百度敏感词.txt", 'r', encoding='gbk')
for line in file:
    res.add(line.strip())
    res.add(pin.get_pinyin(line.strip()).replace("-", ""))

# 读取excel2007文件
wb = load_workbook(filename=r'D:\work_space\敏感词库\网站敏感词库\JasonYSU的敏感词库\敏感词库表统计.xlsx')
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
for rx in range(2, ws.max_row + 1):
    res.add(str(ws.cell(row=rx, column=4).value))
    res.add(pin.get_pinyin(str(ws.cell(row=rx, column=4).value)).replace("-", ""))

# 读取excel2007文件
wb = load_workbook(filename=r'D:\work_space\敏感词库\网站敏感词库\敏感词库汇总版-方石20161027.xlsx')
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
for rx in range(1, ws.max_row + 1):
    temp = str(ws.cell(row=rx, column=1).value).split()
    for item in temp:
        if len(item) >= 2:
            res.add(item)
            res.add(pin.get_pinyin(item).replace("-", ""))
    res.add(str(ws.cell(row=rx, column=1).value))
    res.add(pin.get_pinyin(str(ws.cell(row=rx, column=1).value)).replace("-", ""))

# ---------------------------------------------------------------------------------------- #
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\恶意网站实验室--恶意网址集.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\广告、非法信息类.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\领导名人类.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\论坛需要过滤的不良词语大全.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("*", "")
    temp = temp.replace("=", "")
    temp = temp.replace("http://www.hao123.com/", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\骂人、讽刺类.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\迷信、邪教类.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\其它次要的适合于后台审核的敏感词.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\清理整治网上涉性用品药品非法信息关键词.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\清理整治网上涉性用品药品非法信息关键词.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\色情类.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\涉枪涉爆违法信息关键词.txt", 'r', encoding='utf-8', errors="ignore")
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\时事类.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\高考考生信息.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\网站敏感词1.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\违法信息关键词.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    temp = temp.replace("\"", "")
    temp = temp.split()
    for item in temp:
        res.add(item.strip())
        res.add(pin.get_pinyin(item.strip()).replace("-", ""))
file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\药物毒品类.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))

file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\\2012年最新敏感词列表\政治类.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))

file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\8000多个敏感词过滤 让你的论坛网站安全Words\8000多个敏感词过滤 让你的论坛网站安全Words.txt", 'r', encoding='utf-8', errors="ignore")
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    temp = re.sub(r'["+(+)+\-+=+>+*+]', "", temp)
    temp = temp.split()
    for item in temp:
        res.add(item.strip())
        res.add(pin.get_pinyin(item.strip()).replace("-", ""))

file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\CensorWords\CensorWords.txt", 'r', encoding='gbk')
for line in file:
    if line[0] != "[":
        temp = deepcopy(line)
        temp = temp.replace("={BANNED}", "")
        temp = temp.replace("={MOD}", "")
        res.add(temp.strip())
        res.add(pin.get_pinyin(temp.strip()).replace("-", ""))

file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\敏感词\敏感词库大全.txt", 'r', encoding='gbk')
for line in file:
    res.add(line.replace("|1", "").strip())
    res.add(pin.get_pinyin(line.replace("|1", "").strip()).replace("-", ""))

file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\敏感词1\BBSCH\汇总.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = re.sub(r'\|1', "", temp)
    temp = re.sub(r'\|2', "", temp)
    temp = re.sub(r'\|3', "", temp)
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))

file = open("d:\work_space\敏感词库\网站敏感词库\敏感词库\敏感词2\全部词库.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = re.sub(r'\|1', "", temp)
    temp = re.sub(r'\|2', "", temp)
    temp = re.sub(r'\|3', "", temp)
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))

file = open("D:\work_space\敏感词库\网站敏感词库\敏感词库\敏感词过滤\\01.txt", 'r', encoding='gbk')
for line in file:
    res.add(line.strip())
    res.add(pin.get_pinyin(line.strip()).replace("-", ""))

file = open("D:\work_space\敏感词库\网站敏感词库\敏感词库\敏感词过滤\\02.txt", 'r', encoding='gbk')
for line in file:
    temp = line.split("@")
    for item in temp:
        if item:
            res.add(item.strip())
            res.add(pin.get_pinyin(item.strip()).replace("-", ""))

file = open("D:\work_space\敏感词库\网站敏感词库\敏感词库\敏感词过滤\\CensorWords.txt", 'r', encoding='gbk')
for line in file:
    temp = deepcopy(line)
    temp = temp.replace("={BANNED}", "")
    temp = temp.replace("={MOD}", "")
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))

file = open("D:\work_space\敏感词库\网站敏感词库\敏感词库\敏感词过滤\\CensorWords2.txt", 'r', encoding='gbk')
for line in file:
    if "[" not in line and "]" not in line:
        temp = deepcopy(line)
        temp = temp.replace("={BANNED}", "")
        temp = temp.replace("={MOD}", "")
        temp = temp.replace("=※※", "")
        res.add(temp.strip())
        res.add(pin.get_pinyin(temp.strip()).replace("-", ""))

file = open("D:\work_space\敏感词库\网站敏感词库\CensorWords.txt", 'r', encoding='utf-8', errors='ignore')
for line in file:
    res.add(line.strip())
    res.add(pin.get_pinyin(line.strip()).replace("-", ""))

file = open("D:\work_space\敏感词库\网站敏感词库\敏感词库大全.txt", 'r', encoding='gbk', errors='ignore')
for line in file:
    temp = deepcopy(line)
    temp = re.sub(r'\|1', "", temp)
    temp = re.sub(r'\|2', "", temp)
    temp = re.sub(r'\|3', "", temp)
    res.add(temp.strip())
    res.add(pin.get_pinyin(temp.strip()).replace("-", ""))

res.remove("")
print(len(res))
res.add("射美女")
res.add("操美女")
res.add("狠撸撸")
res.add("撸")
res.add("艹")
res.add("撸射")
res.add("身寸")
res.add("爆射")
res.add("暴射")
res.add("motherfuck")
res.add("雞八")
temp_list = list(res)

for item in temp_list:
    if "www." in str(item):
        res.add(item.replace("www.", "").strip())
    if "{1}" in str(item):
        res.add(item.replace("{1}", "").strip())
    if "{2}" in str(item):
        res.add(item.replace("{2}", "").strip())
    if "{3}" in str(item):
        res.add(item.replace("{3}", "").strip())
    if "{4}" in str(item):
        res.add(item.replace("{4}", "").strip())
one_word = []
# 处理一字词
f = open("D:\work_space\敏感词库\网站敏感词库\\remove.txt", 'r', encoding='gbk')
for line in f:
    if line:
        one_word.append(str(line).strip())
for i in range(len(temp_list)-1 , -1, -1):
    if temp_list[i] not in one_word and len(temp_list[i]) == 1:
        res.remove(temp_list[i])
res.remove("卧槽")
print("阴道" in res)
file = open("CensorWords.txt", 'w', encoding='utf-8')
for line in list(res):
    file.write(str(line).strip() + '\n')

file.close()
